import sys
import os
import pandas as pd
import warnings
import re
import logging
import traceback

# win32com 仅在 Windows 上可用；在 Mac/Linux 上加降级提示，避免 import 直接崩溃
try:
    import win32com.client as win32
except ImportError:
    if sys.platform == "win32":
        # Windows 上找不到 pywin32，提示装依赖
        raise ImportError(
            "本工具需要 pywin32（Windows）。请在终端里运行：\n"
            "    pip install pywin32\n"
            "或者双击项目根目录的 “首次安装.bat”。"
        )
    # 非 Windows 平台：给个 stub，并在真正调用 Excel 时再报错
    print(
        "⚠️  Excel 搜索器依赖 Windows + pywin32 + Excel 应用程序，"
        "在 Mac/Linux 上无法正常使用。\n"
        "请在 Windows 电脑上运行本工具。",
        file=sys.stderr,
    )
    win32 = None  # type: ignore

from openpyxl import load_workbook
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
                             QLineEdit, QPushButton, QComboBox, QTableWidget, QTableWidgetItem, 
                             QFileDialog, QLabel, QRadioButton, QHeaderView, 
                             QMessageBox, QMenu, QTextEdit, QDialog, QFormLayout, QProgressBar)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QSettings
from PyQt6.QtGui import QAction

# 屏蔽警告
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# --- 日志系统 ---
logging.basicConfig(filename='excel_tool_error.log', level=logging.INFO, 
                    format='%(asctime)s - %(levelname)s - %(message)s', encoding='utf-8')

def log_error(msg):
    logging.error(msg)
    logging.error(traceback.format_exc())

# [此处省略你之前提供的 SearchEngine 和 EditorDialog 类，代码保持不变]
# 为了节省篇幅，假设这两个类已经完整包含在这里...

class SearchEngine(QThread):
    progress_signal = pyqtSignal(str)
    result_signal = pyqtSignal(list)
    finished_signal = pyqtSignal(int)

    def __init__(self, folders, conds, base_data=None):
        super().__init__()
        self.folders = [f.strip() for f in folders.split(";") if f.strip()]
        self.conds = conds
        self.base_data = base_data
        self._is_running = True

    def stop(self): self._is_running = False

    def check_match(self, text, op, target):
        if not target: return True
        v, t = str(text).lower(), str(target).lower()
        if op == "包含": return t in v
        if op == "等于": return v == t
        if op == "不包含": return t not in v
        if op == "开头是": return v.startswith(t)
        if op == "结尾是": return v.endswith(t)
        try:
            if op == "大于": return float(v) > float(t)
            if op == "小于": return float(v) < float(t)
        except: return False
        return False

    def run(self):
        try:
            c1, v1, c2, v2, is_and = self.conds
            count = 0
            if self.base_data:
                for item in self.base_data:
                    if not self._is_running: break
                    full_text = item[3]
                    res1 = self.check_match(full_text, c1, v1)
                    match = res1 if c2 == "(无)" else ((res1 and self.check_match(full_text, c2, v2)) if is_and else (res1 or self.check_match(full_text, c2, v2)))
                    if match:
                        self.result_signal.emit(item)
                        count += 1
            else:
                for folder in self.folders:
                    if not self._is_running: break
                    if not os.path.exists(folder): continue
                    for root, _, files in os.walk(folder):
                        for file in files:
                            if not self._is_running: break
                            if file.startswith("~$") or not file.lower().endswith(('.xlsx', '.xls', '.csv', '.xlsm', '.xlsb')): continue
                            file_path = os.path.join(root, file)
                            self.progress_signal.emit(f"🔍 正在检索: {file}")
                            try:
                                if file.lower().endswith('.csv'):
                                    df_dict = {"CSV_Data": pd.read_csv(file_path, encoding='gb18030')}
                                else:
                                    df_dict = pd.read_excel(file_path, sheet_name=None)
                                for sheet_name, df in df_dict.items():
                                    df = df.iloc[:, :26].dropna(axis=1, how='all').fillna("")
                                    headers = df.columns.tolist()
                                    for idx, row in df.iterrows():
                                        row_vals = [str(x).strip() for x in row.values]
                                        full_row_str = " ".join(row_vals)
                                        res1 = self.check_match(full_row_str, c1, v1)
                                        match = res1 if c2 == "(无)" else ((res1 and self.check_match(full_row_str, c2, v2)) if is_and else (res1 or self.check_match(full_row_str, c2, v2)))
                                        if match:
                                            self.result_signal.emit([file, str(sheet_name), str(idx+2), full_row_str, headers, row.tolist(), file_path])
                                            count += 1
                            except Exception as e:
                                log_error(f"读取失败 {file}: {e}")
            self.finished_signal.emit(count)
        except Exception as e:
            log_error(f"引擎异常: {e}")

class EditorDialog(QDialog):
    def __init__(self, headers, values, parent=None):
        super().__init__(parent)
        self.setWindowTitle("📝 在线编辑行数据")
        self.setMinimumWidth(500)
        layout = QFormLayout(self)
        self.inputs = []
        for h, v in zip(headers, values):
            le = QLineEdit(str(v))
            layout.addRow(QLabel(f"{h}:"), le)
            self.inputs.append(le)
        self.btn = QPushButton("💾 确认保存修改")
        self.btn.clicked.connect(self.accept)
        layout.addRow(self.btn)

    def get_new_values(self):
        return [i.text() for i in self.inputs]

# --- 这里是 UI 主类 ExcelApp ---
class ExcelApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Excel/WPS 全能搜索工具 V3.1")
        self.resize(1300, 900)
        self.settings = QSettings("MyStudio", "SearchProV3")
        self.undo_stack, self.active_filters = [], []
        self.is_busy = False
        self.init_ui()
        self.load_history()

    # [此处省略你代码中的所有其他方法，如 init_ui, new_search, render_row 等...]
    # 请务必保留你原来的完整逻辑代码
    
    # ... (此处补全你提供的所有函数) ...
    def init_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)
        nav = QHBoxLayout()
        self.path_input = QComboBox(); self.path_input.setEditable(True); self.path_input.setPlaceholderText("多文件夹请用分号 ; 隔开")
        btn_browse = QPushButton("📁 浏览添加"); btn_browse.clicked.connect(self.on_browse)
        self.history_combo = QComboBox(); self.history_combo.setPlaceholderText("关键词历史"); self.history_combo.activated.connect(lambda: self.val_1.setText(self.history_combo.currentText()))
        nav.addWidget(QLabel("搜索路径:"), 1); nav.addWidget(self.path_input, 5); nav.addWidget(btn_browse, 1); nav.addWidget(QLabel("关键词历史:"), 1); nav.addWidget(self.history_combo, 2)
        main_layout.addLayout(nav)
        cond_box = QHBoxLayout()
        self.cond_1 = QComboBox(); self.cond_1.addItems(["包含", "等于", "不包含", "开头是", "结尾是", "大于", "小于"])
        self.val_1 = QLineEdit(); self.radio_and = QRadioButton("与"); self.radio_and.setChecked(True)
        self.radio_or = QRadioButton("或"); self.val_2 = QLineEdit()
        self.cond_2 = QComboBox(); self.cond_2.addItems(["(无)", "包含", "等于", "不包含", "开头是", "结尾是", "大于", "小于"])
        for w in [self.cond_1, self.val_1, self.radio_and, self.radio_or, self.cond_2, self.val_2]: cond_box.addWidget(w)
        main_layout.addLayout(cond_box)
        btn_box = QHBoxLayout()
        self.btn_run = QPushButton("🚀 开始全面搜索"); self.btn_run.clicked.connect(self.new_search)
        self.btn_refine = QPushButton("🔍 结果内二次筛选"); self.btn_refine.clicked.connect(self.refine_search)
        self.btn_back = QPushButton("⬅ 后退"); self.btn_back.clicked.connect(self.undo_action)
        self.btn_stop = QPushButton("🛑 停止"); self.btn_stop.setEnabled(False); self.btn_stop.clicked.connect(self.stop_search)
        self.btn_run.setStyleSheet("background-color: #217346; color: white; font-weight: bold; height: 35px;")
        for b in [self.btn_run, self.btn_refine, self.btn_back, self.btn_stop]: btn_box.addWidget(b)
        main_layout.addLayout(btn_box)
        self.pbar = QProgressBar(); self.pbar.setVisible(False); main_layout.addWidget(self.pbar)
        self.table = QTableWidget(0, 4); self.table.setHorizontalHeaderLabels(["文件名", "工作表", "行号", "详情内容 (右键操作)"]); self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch); self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu); self.table.customContextMenuRequested.connect(self.on_context_menu); main_layout.addWidget(self.table)
        self.status = QLabel("就绪"); main_layout.addWidget(self.status)

    def on_browse(self):
        p = QFileDialog.getExistingDirectory()
        if p:
            current = self.path_input.currentText()
            self.path_input.setCurrentText(f"{current};{p}" if current and p not in current else p)
    
    def stop_search(self):
        if hasattr(self, 'worker'): self.worker.stop()
        self.set_loading(False)

    def set_loading(self, busy):
        self.is_busy = busy
        self.btn_run.setEnabled(not busy); self.btn_refine.setEnabled(not busy); self.btn_stop.setEnabled(busy); self.pbar.setVisible(busy)
        if busy: self.pbar.setRange(0, 0)

    def new_search(self):
        if self.is_busy: return
        self.undo_stack.clear(); self.active_filters = []; self.execute_search(refine=False)

    def refine_search(self):
        if self.is_busy or self.table.rowCount() == 0: return
        self.undo_stack.append(self.get_current_state()); self.execute_search(refine=True)

    def execute_search(self, refine=False):
        folders = self.path_input.currentText()
        if not refine and not folders.strip(): return
        self.set_loading(True); self.table.setRowCount(0)
        kws = [k for k in [self.val_1.text(), self.val_2.text()] if k]
        self.active_filters.append({"kws": kws, "ops": [self.cond_1.currentText(), self.cond_2.currentText()]})
        base = self.undo_stack[-1]["data"] if refine else None
        conds = [self.cond_1.currentText(), self.val_1.text(), self.cond_2.currentText(), self.val_2.text(), self.radio_and.isChecked()]
        self.worker = SearchEngine(folders, conds, base); self.worker.progress_signal.connect(self.status.setText); self.worker.result_signal.connect(self.render_row); self.worker.finished_signal.connect(self.on_done); self.worker.start()

    def render_row(self, data):
        r = self.table.rowCount(); self.table.insertRow(r)
        self.table.setItem(r, 0, QTableWidgetItem(data[0])); self.table.setItem(r, 1, QTableWidgetItem(data[1])); self.table.setItem(r, 2, QTableWidgetItem(data[2]))
        edit = QTextEdit(); edit.setReadOnly(True); edit.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        hdrs, vals = data[4], data[5]
        html_parts = []
        colors = ["red", "#005fb8", "#228b22", "#ff8c00", "#8b008b"]
        for h, v in zip(hdrs, vals):
            v_s = str(v).strip()
            if not v_s: continue
            for idx, flt in enumerate(self.active_filters):
                color = colors[idx % len(colors)]
                for kw in flt["kws"]:
                    if kw: v_s = re.sub(f"({re.escape(kw)})", rf'<span style="color:{color}; font-weight:bold;">\1</span>', v_s, flags=re.IGNORECASE)
            html_parts.append(f"<b>[{h}]:</b> {v_s}")
        edit.setHtml(" &nbsp; | &nbsp; ".join(html_parts))
        metadata = {"path": data[6], "sheet": data[1], "rno": data[2], "hdrs": hdrs, "vals": vals, "plain": data[3]}
        for k, v in metadata.items(): edit.setProperty(k, v)
        self.table.setCellWidget(r, 3, edit); self.table.setRowHeight(r, 65)

    def on_done(self, count):
        self.set_loading(False); self.status.setText(f"✅ 发现 {count} 条结果")
        path_text, kw_text = self.path_input.currentText(), self.val_1.text()
        if path_text:
            paths = self.settings.value("paths", [])
            if path_text in paths: paths.remove(path_text)
            self.settings.setValue("paths", ([path_text] + paths)[:15])
        if kw_text:
            kws = self.settings.value("keywords", [])
            if kw_text in kws: kws.remove(kw_text)
            self.settings.setValue("keywords", ([kw_text] + kws)[:20])
        self.load_history()

    def load_history(self):
        self.path_input.blockSignals(True); self.path_input.clear(); self.path_input.addItems(self.settings.value("paths", [])); self.path_input.setCurrentIndex(-1); self.path_input.blockSignals(False)
        self.history_combo.clear(); self.history_combo.addItems(self.settings.value("keywords", []))

    def on_context_menu(self, pos):
        if not self.table.itemAt(pos): return
        menu = QMenu(); a1 = QAction("📍 在 Excel 中打开并定位行", self); a2 = QAction("📝 快速编辑此行 (物理修改)", self)
        a1.triggered.connect(self.locate_in_excel); a2.triggered.connect(self.inline_edit); menu.addActions([a1, a2]); menu.exec(self.table.mapToGlobal(pos))

    def locate_in_excel(self):
        w = self.table.cellWidget(self.table.currentRow(), 3)
        try:
            try: excel = win32.GetActiveObject("Excel.Application")
            except: excel = win32.Dispatch("Excel.Application")
            excel.Visible = True; wb = excel.Workbooks.Open(os.path.abspath(w.property("path"))); ws = wb.Worksheets(w.property("sheet")); ws.Activate()
            r = int(w.property("rno")); excel.ActiveWindow.ScrollRow = max(1, r - 5); ws.Rows(r).Select()
        except Exception as e: QMessageBox.critical(self, "错误", str(e))

    def inline_edit(self):
        idx = self.table.currentRow(); w = self.table.cellWidget(idx, 3)
        if not w: return
        path, sheet, rno, hdrs, old_vals = w.property("path"), w.property("sheet"), int(w.property("rno")), w.property("hdrs"), w.property("vals")
        dlg = EditorDialog(hdrs, old_vals, self)
        if dlg.exec():
            new_vals = dlg.get_new_values()
            try:
                wb = load_workbook(path); ws = wb[sheet]
                for c, v in enumerate(new_vals, 1): ws.cell(row=rno, column=c).value = v
                wb.save(path); w.setProperty("vals", new_vals)
                html_parts = []
                colors = ["red", "#005fb8", "#228b22", "#ff8c00", "#8b008b"]
                for h, v in zip(hdrs, new_vals):
                    v_s = str(v).strip()
                    if not v_s: continue
                    for f_idx, flt in enumerate(self.active_filters):
                        color = colors[f_idx % len(colors)]
                        for kw in flt["kws"]:
                            if kw: v_s = re.sub(f"({re.escape(kw)})", rf'<span style="color:{color}; font-weight:bold;">\1</span>', v_s, flags=re.IGNORECASE)
                    html_parts.append(f"<b>[{h}]:</b> {v_s}")
                w.setHtml(" &nbsp; | &nbsp; ".join(html_parts)); self.status.setText(f"✅ 第 {rno} 行已物理同步并刷新")
            except Exception as e: QMessageBox.critical(self, "错误", f"保存失败: {e}")

    def get_current_state(self):
        data = []
        for r in range(self.table.rowCount()):
            w = self.table.cellWidget(r, 3)
            data.append([self.table.item(r,0).text(), self.table.item(r,1).text(), self.table.item(r,2).text(), w.property("plain"), w.property("hdrs"), w.property("vals"), w.property("path")])
        return {"data": data, "filters": list(self.active_filters)}

    def undo_action(self):
        if self.undo_stack:
            state = self.undo_stack.pop(); self.table.setRowCount(0); self.active_filters = state["filters"]
            for d in state["data"]: self.render_row(d)

# 标准入口：确保脚本可以被直接运行
if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = ExcelApp()
    window.show()
    sys.exit(app.exec())