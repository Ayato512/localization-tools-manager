import os
import json
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook, Workbook

CONFIG_FILE = "excel_merger_config.json"

class ExcelMergerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel 多表原样合并工具 v1.0")
        self.root.geometry("620x430")
        self.root.resizable(False, False)
        
        # 初始化变量
        self.input_files = []
        self.last_input_dir = ""
        self.last_output_dir = ""
        
        # 加载历史记忆路径
        self.load_config()
        
        # 界面布局
        self.create_widgets()

    def load_config(self):
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                    config = json.load(f)
                    self.last_input_dir = config.get("last_input_dir", "")
                    self.last_output_dir = config.get("last_output_dir", "")
            except Exception:
                pass

    def save_config(self):
        config = {
            "last_input_dir": self.last_input_dir,
            "last_output_dir": self.last_output_dir
        }
        try:
            with open(CONFIG_FILE, "w", encoding="utf-8") as f:
                json.dump(config, f, ensure_ascii=False, indent=4)
        except Exception:
            pass

    def create_widgets(self):
        padding = {"padx": 15, "pady": 8}
        
        # 1. 输入文件选择
        input_frame = tk.LabelFrame(self.root, text=" 1. 选择待合并的 Excel 文件 (支持多选) ")
        input_frame.pack(fill="x", **padding)
        
        self.btn_select_files = tk.Button(input_frame, text="浏览并选择文件", command=self.select_files, width=15)
        self.btn_select_files.pack(side="left", padx=10, pady=10)
        
        self.lbl_file_count = tk.Label(input_frame, text="未选择任何文件", fg="gray")
        self.lbl_file_count.pack(side="left", padx=10, pady=10)

        # 文件列表预览
        list_frame = tk.Frame(self.root)
        list_frame.pack(fill="both", expand=True, padx=15)
        
        self.file_listbox = tk.Listbox(list_frame, height=6, selectmode=tk.EXTENDED)
        self.file_listbox.pack(side="left", fill="both", expand=True)
        
        scrollbar = tk.Scrollbar(list_frame, orient="vertical", command=self.file_listbox.yview)
        scrollbar.pack(side="right", fill="y")
        self.file_listbox.config(yscrollcommand=scrollbar.set)

        # 2. 输出路径选择
        output_frame = tk.LabelFrame(self.root, text=" 2. 设置输出文件路径与名称 ")
        output_frame.pack(fill="x", **padding)
        
        self.btn_select_output = tk.Button(output_frame, text="设置输出文件", command=self.select_output_file, width=15)
        self.btn_select_output.pack(side="left", padx=10, pady=10)
        
        self.entry_output_path = tk.Entry(output_frame, width=45)
        self.entry_output_path.pack(side="left", padx=10, pady=10, fill="x", expand=True)

        # 3. 执行按钮
        btn_frame = tk.Frame(self.root)
        btn_frame.pack(fill="x", padx=15, pady=15)
        
        self.btn_run = tk.Button(btn_frame, text="开始按 A-Z 排序合并", command=self.merge_excels, bg="#107c41", fg="white", font=("Microsoft YaHei", 10, "bold"), height=2)
        self.btn_run.pack(fill="x")

    def select_files(self):
        initial_dir = self.last_input_dir if os.path.exists(self.last_input_dir) else os.getcwd()
        files = filedialog.askopenfilenames(
            title="选择 Excel 文件",
            initialdir=initial_dir,
            filetypes=[("Excel Files", "*.xlsx")]
        )
        if files:
            self.input_files = list(files)
            self.last_input_dir = os.path.dirname(self.input_files[0])
            self.save_config()
            
            # 更新界面显示
            self.file_listbox.delete(0, tk.END)
            for f in sorted(self.input_files, key=lambda x: os.path.basename(x).lower()):
                self.file_listbox.insert(tk.END, os.path.basename(f))
            self.lbl_file_count.config(text=f"已选择 {len(self.input_files)} 个文件 (已按名称自动排序)", fg="green")

    def select_output_file(self):
        initial_dir = self.last_output_dir if os.path.exists(self.last_output_dir) else (self.last_input_dir if os.path.exists(self.last_input_dir) else os.getcwd())
        output_file = filedialog.asksaveasfilename(
            title="保存合并后的文件",
            initialdir=initial_dir,
            initialfile="合并结果.xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            defaultextension=".xlsx"
        )
        if output_file:
            self.entry_output_path.delete(0, tk.END)
            self.entry_output_path.insert(0, output_file)
            self.last_output_dir = os.path.dirname(output_file)
            self.save_config()

    def merge_excels(self):
        if not self.input_files:
            messagebox.showwarning("警告", "请先选择需要合并的 Excel 文件！")
            return
        
        output_path = self.entry_output_path.get().strip()
        if not output_path:
            messagebox.showwarning("警告", "请设置输出文件的保存路径及名称！")
            return

        # 1. 严格按文件名 A-Z 排序 (忽略大小写)
        sorted_files = sorted(self.input_files, key=lambda x: os.path.basename(x).lower())
        
        # 2. 第一轮扫描：检测全局工作表（Sheet）名称冲突
        sheet_name_map = {}  # sheet_name -> file_path
        duplicate_sheets = []

        for file_path in sorted_files:
            try:
                wb = load_workbook(file_path, read_only=True)
                for sheet_name in wb.sheetnames:
                    if sheet_name in sheet_name_map:
                        duplicate_sheets.append((sheet_name, sheet_name_map[sheet_name], file_path))
                    else:
                        sheet_name_map[sheet_name] = file_path
                wb.close()
            except Exception as e:
                messagebox.showerror("错误", f"读取文件异常: {os.path.basename(file_path)}\n{str(e)}")
                return

        # 3. 冲突报错中断
        if duplicate_sheets:
            error_msg = "检测到以下文件存在同名工作表（Sheet），合并强行终止：\n\n"
            for sheet, file1, file2 in duplicate_sheets[:5]:  # 最多显示5条避免弹窗过长
                error_msg += f"• Sheet名称: [{sheet}]\n  冲突文件1: {os.path.basename(file1)}\n  冲突文件2: {os.path.basename(file2)}\n\n"
            if len(duplicate_sheets) > 5:
                error_msg += f"...以及其他 {len(duplicate_sheets) - 5} 处冲突。\n\n"
            error_msg += "请自行修改原 Excel 中的工作表名称后重试。"
            messagebox.showerror("Sheet 同名冲突", error_msg)
            return

        # 4. 开始原样合并 (保留公式与样式)
        try:
            new_wb = Workbook()
            # 默认会带一个Sheet，先留着，后面删掉
            default_sheet = new_wb.active
            
            for file_path in sorted_files:
                # 不使用 data_only=True，确保保留原公式
                src_wb = load_workbook(file_path, data_only=False, keep_vba=True)
                
                for sheet_name in src_wb.sheetnames:
                    src_sheet = src_wb[sheet_name]
                    # 创建新 sheet 并复制基础属性
                    new_sheet = new_wb.create_sheet(title=sheet_name)
                    
                    # 复制视图及网格线属性
                    new_sheet.views.sheetView[0].showGridLines = src_sheet.views.sheetView[0].showGridLines
                    
                    # 遍历复制单元格、值、样式和公式
                    for row in src_sheet.iter_rows():
                        for cell in row:
                            new_cell = new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                            
                            # 复制样式 (规避 StyleProxy 导致的 unhashable 错误)
                            if cell.has_style:
                                from copy import copy
                                try:
                                    new_cell.font = copy(cell.font)
                                    new_cell.border = copy(cell.border)
                                    new_cell.fill = copy(cell.fill)
                                    new_cell.alignment = copy(cell.alignment)
                                    new_cell.number_format = cell.number_format
                                    new_cell.protection = copy(cell.protection)
                                except Exception:
                                    # 如果个别极端样式依然报错，作为保底方案直接赋文本值，不影响合并中断
                                    pass
                    
                    # 复制行高
                    for row_idx, row_dim in src_sheet.row_dimensions.items():
                        new_sheet.row_dimensions[row_idx].height = row_dim.height
                        
                    # 复制列宽
                    for col_letter, col_dim in src_sheet.column_dimensions.items():
                        new_sheet.column_dimensions[col_letter].width = col_dim.width
                        
                    # 复制合并单元格
                    for merged_range in src_sheet.merged_cells.ranges:
                        new_sheet.merge_cells(str(merged_range))
                        
                src_wb.close()

            # 移除最开始创建的默认空白 Sheet
            if "Sheet" in new_wb.sheetnames and len(new_wb.sheetnames) > 1:
                new_wb.remove(default_sheet)

            # 保存最终结果
            new_wb.save(output_path)
            new_wb.close()
            
            messagebox.showinfo("成功", f"所有文件已成功按 A-Z 顺序合并完成！\n输出路径：{output_path}")
            
        except Exception as e:
            messagebox.showerror("合并失败", f"在合并过程中发生未预期的错误：\n{str(e)}")

if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelMergerApp(root)
    root.mainloop()
