import sys
import os
import re
import openpyxl
from copy import copy
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                             QPushButton, QFileDialog, QLabel, QTextEdit, QHBoxLayout)
from PyQt6.QtCore import Qt, QThread, pyqtSignal

# --- 核心逻辑类 ---
class ExcelProcessor:
    @staticmethod
    def split_excel(file_path, output_dir):
        """拆分逻辑：保留原格式、图片、公式"""
        wb_template = openpyxl.load_workbook(file_path, data_only=False)
        all_sheets = wb_template.sheetnames
        
        for target_sheet in all_sheets:
            # 重新加载以确保操作的是完整原件
            wb_current = openpyxl.load_workbook(file_path)
            for sheet_name in wb_current.sheetnames:
                if sheet_name != target_sheet:
                    del wb_current[sheet_name]
            
            clean_name = "".join([c for c in target_sheet if c.isalnum() or c in (' ', '_', '-')]).strip()
            output_path = os.path.join(output_dir, f"{clean_name}.xlsx")
            wb_current.save(output_path)
            yield f"已导出: {clean_name}.xlsx"

    @staticmethod
    def merge_excels(input_dir, output_file):
        """合并逻辑：尽可能还原格式与图片"""
        combined_wb = openpyxl.Workbook()
        default_sheet = combined_wb.active
        combined_wb.remove(default_sheet)

        files = [f for f in os.listdir(input_dir) if f.endswith('.xlsx') and not f.startswith('~$')]
        files.sort()

        for file_name in files:
            file_path = os.path.join(input_dir, file_name)
            source_wb = openpyxl.load_workbook(file_path)
            
            for sheet_name in source_wb.sheetnames:
                source_sheet = source_wb[sheet_name]
                target_sheet = combined_wb.create_sheet(title=sheet_name)
                
                # 复制列宽
                for col_name, col_dim in source_sheet.column_dimensions.items():
                    target_sheet.column_dimensions[col_name].width = col_dim.width
                
                # 复制行高、内容与样式
                for row in source_sheet.iter_rows():
                    for cell in row:
                        new_cell = target_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                        if cell.has_style:
                            new_cell.font = copy(cell.font)
                            new_cell.border = copy(cell.border)
                            new_cell.fill = copy(cell.fill)
                            new_cell.number_format = copy(cell.number_format)
                            new_cell.alignment = copy(cell.alignment)

                # 复制合并单元格
                for merged_range in source_sheet.merged_cells.ranges:
                    target_sheet.merge_cells(str(merged_range))

                # 复制图片
                if hasattr(source_sheet, '_images'):
                    from copy import deepcopy
                    for img in source_sheet._images:
                        new_img = deepcopy(img)
                        target_sheet.add_image(new_img)
            
            yield f"已合并: {file_name}"
        
        combined_wb.save(output_file)

# --- PyQt6 界面类 ---
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Excel 本地化助手 - 拆分与整合工具")
        self.setMinimumSize(600, 400)
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()

        self.label = QLabel("选择操作并按照提示选择文件或文件夹")
        self.label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.label)

        self.log_view = QTextEdit()
        self.log_view.setReadOnly(True)
        self.log_view.setPlaceholderText("操作日志将显示在这里...")
        layout.addWidget(self.log_view)

        btn_layout = QHBoxLayout()
        self.btn_split = QPushButton("拆分 Excel (保留格式/图片)")
        self.btn_split.clicked.connect(self.handle_split)
        
        self.btn_merge = QPushButton("合并文件夹 (还原格式/图片)")
        self.btn_merge.clicked.connect(self.handle_merge)

        btn_layout.addWidget(self.btn_split)
        btn_layout.addWidget(self.btn_merge)
        layout.addLayout(btn_layout)

        container = QWidget()
        container.setLayout(layout)
        self.setCentralWidget(container)

    def log(self, message):
        self.log_view.append(message)

    def handle_split(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "选择要拆分的 Excel", "", "Excel Files (*.xlsx)")
        if file_path:
            output_dir = QFileDialog.getExistingDirectory(self, "选择输出文件夹")
            if output_dir:
                self.log(f"--- 开始拆分: {os.path.basename(file_path)} ---")
                for msg in ExcelProcessor.split_excel(file_path, output_dir):
                    self.log(msg)
                self.log("✅ 拆分任务完成！")

    def handle_merge(self):
        input_dir = QFileDialog.getExistingDirectory(self, "选择包含多个子表的文件夹")
        if input_dir:
            output_file, _ = QFileDialog.getSaveFileName(self, "保存整合后的文件", "整合结果.xlsx", "Excel Files (*.xlsx)")
            if output_file:
                self.log(f"--- 开始整合文件夹: {os.path.basename(input_dir)} ---")
                try:
                    for msg in ExcelProcessor.merge_excels(input_dir, output_file):
                        self.log(msg)
                    self.log(f"✅ 整合完成！输出至: {output_file}")
                except Exception as e:
                    self.log(f"❌ 错误: {str(e)}")

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())